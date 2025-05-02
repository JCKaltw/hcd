# version1a.py
import os
import pandas as pd
import matplotlib.pyplot as plt
from datetime import timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from google.colab import drive

# Mount Google Drive
if not os.path.exists('/content/drive'):
    drive.mount('/content/drive', force_remount=True)
    print('Drive mounted')
else:
    print('Drive already mounted at /content/drive')

# Configuration
source_folder = "/content/drive/My Drive/test"
target_folder = "/content/drive/My Drive/test_done"
os.makedirs(target_folder, exist_ok=True)
source_folder = "/content/drive/My Drive/test"
target_folder = "/content/drive/My Drive/test_done"

def process_file(filepath, savepath):
    xls = pd.ExcelFile(filepath)
    raw_df = pd.read_excel(xls, sheet_name=0, header=None)

    # Detect header
    header_row_idx = raw_df[raw_df.apply(lambda row: row.astype(str).str.contains("State").any(), axis=1)].index[0]
    df = pd.read_excel(xls, sheet_name=0, header=header_row_idx).dropna(axis=1, how='all')
    original_df = pd.read_excel(xls, sheet_name=0, header=None)

    # Rename 7th column to "Note"
    if df.columns.size >= 7:
        print(f"📄 Processing file: {filepath}")
        df.columns.values[6] = "Note"
        if "Note" in df.columns:
            df = df[df["Note"] == "Test Run"].copy()
        else:
            print(f"❌ 'Note' column not found after renaming in: {filepath}")
            return
    else:
        print(f"❌ Not enough columns to rename 7th column to 'Note' in: {filepath}")
        return

    # Insert "Device Name" and "MAC Serial #"
    device_name = raw_df.iloc[0, 1]
    mac_serial = str(raw_df.iloc[1, 0]).replace("DevID: ", "")
    df["Device Name"] = device_name
    df["MAC Serial #"] = mac_serial
    cols = ["Device Name", "MAC Serial #"] + [col for col in df.columns if col not in ["Device Name", "MAC Serial #"]]
    df = df[cols]

    # Add Enable/Disable columns
    df["Enable"] = df["State"].apply(lambda x: 1 if x == "Enable" else "")
    df["Disable"] = df["State"].apply(lambda x: 1 if x == "Disable" else "")

    # Step 4: Clean up duplicate/missing timestamps before heat detection
    df["Date"] = pd.to_datetime(df["Date"])
    df = df.sort_values("Date").reset_index(drop=True)

    # Capture duplicate rows (keep last, discard others)
    duplicates = df[df.duplicated(subset="Date", keep="last")]

    # Remove duplicates by keeping the last occurrence
    df = df.drop_duplicates(subset="Date", keep="last")

    # Generate a complete timestamp range at 1-minute frequency
    full_range = pd.date_range(start=df["Date"].min(), end=df["Date"].max(), freq="1min")
    df_full = pd.DataFrame({"Date": full_range})

    # Merge to find missing times
    df = pd.merge(df_full, df, on="Date", how="left")

    # Forward fill row values from previous valid row
    fill_cols = ["Device Name", "MAC Serial #", "Enable", "Disable", "Note"]
    df[fill_cols] = df[fill_cols].fillna(method="ffill")

    # Save discarded rows
    discarded = pd.concat([duplicates, df[df.isnull().any(axis=1)]], ignore_index=True)
    df = df.dropna(subset=["State", "Supply Temp/C", "Return Temp/C"]).copy()

    # Continue with filtered data
    df_filtered = df.copy()
    df_filtered["Date"] = pd.to_datetime(df_filtered["Date"])
    supply = df_filtered["Supply Temp/C"].values
    return_temp = df_filtered["Return Temp/C"].values

    heating_state, heating_group, in_heating, group_id = [], [], False, 0
    triggered_rows = set()
    for i in range(len(supply)):
        if i == 0:
            heating_state.append("Off")
            heating_group.append(0)
            continue
        delta = supply[i] - supply[i - 1]
        if not in_heating and delta > 5 and supply[i] > return_temp[i] and i not in triggered_rows:
            in_heating = True
            group_id += 1
            heating_state.append("On")
            heating_group.append(group_id)
            triggered_rows.add(i)
        elif in_heating and delta <= -2.7:
            in_heating = False
            heating_state.append("Off")
            heating_group.append(0)
        else:
            heating_state.append("On" if in_heating else "Off")
            heating_group.append(group_id if in_heating else 0)

    df_filtered["Heating"] = heating_state
    df_filtered["Heating_Group"] = heating_group

    # Optional: helpful for debugging
    df_filtered["Delta Supply"] = df_filtered["Supply Temp/C"].diff()
    df_filtered["Delta Debug"] = df_filtered["Delta Supply"].apply(lambda x: f"{x:.2f}" if pd.notnull(x) else "")
    df_filtered["Trigger Debug"] = [
        "TRIGGERED" if df_filtered.loc[i, "Delta Supply"] > 5 else ""
        for i in range(len(df_filtered))
    ]

    # ✅ Validate heating groups: at least 60% of rows with Supply ≥ Return + 7°C
    valid_heating = [
        gid for gid in df_filtered["Heating_Group"].unique()
        if gid != 0 and
        (df_filtered[df_filtered["Heating_Group"] == gid]["Supply Temp/C"] >=
         df_filtered[df_filtered["Heating_Group"] == gid]["Return Temp/C"] + 7).mean() >= 0.6
    ]
    df_filtered.loc[~df_filtered["Heating_Group"].isin(valid_heating), ["Heating", "Heating_Group"]] = ["Off", 0]

    # Extract heating group on/off times
    group_times = df_filtered[df_filtered["Heating_Group"] > 0].groupby("Heating_Group")["Date"].agg(["min", "max"]).reset_index()
    group_times.columns = ["Heating_Group", "Heat On", "Heat Off"]
    df_filtered = df_filtered.merge(group_times, on="Heating_Group", how="left")

    # Step 6: Keep full hours where heating occurred
    df_filtered["Hour"] = df_filtered["Date"].dt.floor("h")
    heat_hours = df_filtered[df_filtered["Heating"] == "On"]["Hour"].unique()
    df["Date"] = pd.to_datetime(df["Date"])
    df = pd.merge(df, df_filtered[["Date", "Heating", "Heating_Group"]], on="Date", how="left")
    df["Heating"] = df["Heating"].fillna("Off")
    df["Heating_Group"] = df["Heating_Group"].fillna(0).astype(int)
    df["Hour"] = df["Date"].dt.floor("h")
    heat_data_set = df[df["Hour"].isin(heat_hours)].copy()

    # Step 7: Summarize hours with >=55 mins consistent Enable/Disable
    summaries = []
    for hour in sorted(set(heat_data_set["Date"].dt.floor("h"))):
        hour_data = heat_data_set[heat_data_set["Date"].dt.floor("h") == hour]
        enable_count = pd.to_numeric(hour_data["Enable"], errors='coerce').fillna(0).astype(int).sum() if "Enable" in hour_data else 0
        disable_count = pd.to_numeric(hour_data["Disable"], errors='coerce').fillna(0).astype(int).sum() if "Disable" in hour_data else 0
        if enable_count >= 55 or disable_count >= 55:
            summaries.append({
                "Device Name": hour_data["Device Name"].iloc[0],
                "MAC Serial #": hour_data["MAC Serial #"].iloc[0],
                "Date/Time On": hour,
                "Date/Time Off": hour + timedelta(minutes=59),
                "Enable": 1 if enable_count >= 55 else "",
                "Disable": 1 if disable_count >= 55 else "",
                "Heating On": (hour_data["Heating"] == "On").sum()
            })

    summary_df = pd.DataFrame(summaries)
    print("Summary Rows:", len(summary_df))

    # Save outputs
    with pd.ExcelWriter(savepath, engine='openpyxl') as writer:
        original_df.to_excel(writer, sheet_name="Original Data", index=False)
        df.to_excel(writer, sheet_name="Filtered Test Run", index=False)
        heat_data_set.to_excel(writer, sheet_name="Heating Data Set", index=False)
        summary_df.to_excel(writer, sheet_name="Heat Cleaned Data", index=False)
        discarded.to_excel(writer, sheet_name="Discarded", index=False)
    print(f"✅ Processed and saved: {savepath}")

    # Reopen workbook to apply highlighting
    wb = load_workbook(savepath)
    ws = wb["Filtered Test Run"]

    # Define fill
    light_orange_fill = PatternFill(start_color="FFD8B1", end_color="FFD8B1", fill_type="solid")

    # Find column indexes
    headers = {cell.value: cell.column_letter for cell in ws[1]}
    supply_col = headers.get("Supply Temp/C")
    heating_col = headers.get("Heating")

    # Highlight cells where Heating == "On"
    for row in range(2, ws.max_row + 1):
        if ws[f"{heating_col}{row}"].value == "On":
            if supply_col:
                ws[f"{supply_col}{row}"].fill = light_orange_fill
            if heating_col:
                ws[f"{heating_col}{row}"].fill = light_orange_fill

    # Save highlighted file
    wb.save(savepath)

# Process all Excel files
for filename in os.listdir(source_folder):
    if filename.endswith(".xlsx") and not filename.startswith("~$"):
        full_path = os.path.join(source_folder, filename)
        name, ext = os.path.splitext(filename)
        save_path = os.path.join(target_folder, f"{name}_heat min per hour.xlsx")
        process_file(full_path, save_path)
