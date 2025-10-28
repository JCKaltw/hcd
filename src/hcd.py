# /Users/chris/projects/heat-cycle-detection/src/hcd.py

import os
import sys
import pandas as pd
import matplotlib.pyplot as plt
from datetime import timedelta, datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill  # <-- Added from version1a
# from google.colab import drive  # Removed for local usage

# --------------------------------------------------------------------------------
# The following lines were removed to eliminate Google Colab references.
# --------------------------------------------------------------------------------

# Configuration
source_folder = "./test"        # You can modify as needed (overridden below in main())
target_folder = "./test_done"   # You can modify as needed

# Ensure that target_folder exists
os.makedirs(target_folder, exist_ok=True)

# --------------------------------------------------------------------------------
# Imports for PostgreSQL insertion and JSON output
import psycopg2
import pytz
import json

# --------------------------------------------------------------------------------

def hex_upper(serial: str) -> str:
    """
    Normalize MAC serial number to consistent format:
    - Lowercase 0x prefix (if already present)
    - Uppercase hex digits
    - Do NOT add prefix if it doesn't exist

    Examples:
        a34f -> A34F
        0Xa34f -> 0xA34F
        0xa34f -> 0xA34F
        B0A732E4DA4A -> B0A732E4DA4A
        b0a732e61eba -> B0A732E61EBA

    Args:
        serial: Raw serial number string

    Returns:
        Normalized serial number (with or without 0x prefix based on input)
    """
    if not serial or not isinstance(serial, str):
        return serial

    # Strip whitespace
    serial = serial.strip()

    # Check for 0x or 0X prefix
    has_prefix = serial.lower().startswith('0x')

    if has_prefix:
        # Has prefix: normalize to lowercase 0x + uppercase hex
        hex_part = serial[2:].upper()
        return f"0x{hex_part}"
    else:
        # No prefix: just uppercase the hex digits
        return serial.upper()


def connect_to_postgres():
    """
    Connect to the PostgreSQL database using credentials from environment variables
    and the .pgpass file for the password, as requested.
    """
    return psycopg2.connect(
        host=os.getenv('PGHOST_2'),
        database=os.getenv('PGDATABASE_2'),
        user=os.getenv('PGUSER_2'),
        password=None,  # psycopg2 will automatically use the password from ~/.pgpass
        port=os.getenv('PGPORT_2')
    )


def process_file(filepath, savepath, insert_db=False, do_upserts=False, dry_run=False):
    # Initialize counters for this file
    summary_rows_count = 0
    heating_devices_count = 0
    heating_device_readings_count = 0
    # Initialize list to store heating device info for JSON output
    heating_serial_devices = []
    # Initialize device statistics dictionary
    device_stats = {
        'filepath': filepath,
        'device_name': None,
        'device_serial': None,
        'test_run_rows': 0,
        'supply_min': None,
        'supply_max': None,
        'supply_mean': None,
        'return_min': None,
        'return_max': None,
        'return_mean': None,
        'diff_min': None,
        'diff_max': None,
        'diff_mean': None,
        'rows_above_7c_threshold': 0,
        'rows_supply_gt_return': 0,
        'heating_groups_detected': 0,
        'valid_heating_groups': 0,
        'summary_rows': 0,
        'status': 'unknown'
    }

    xls = pd.ExcelFile(filepath)
    raw_df = pd.read_excel(xls, sheet_name=0, header=None)

    # Detect header
    header_row_idx = raw_df[raw_df.apply(lambda row: row.astype(str).str.contains("State").any(), axis=1)].index[0]
    df = pd.read_excel(xls, sheet_name=0, header=header_row_idx).dropna(axis=1, how='all')
    original_df = pd.read_excel(xls, sheet_name=0, header=None)

    # Rename 7th column to "Note"
    if df.columns.size >= 7:
        print(f"📄 Processing file: {filepath}")
        df.columns.values[6] = "Note"
        if "Note" in df.columns:
            df = df[df["Note"] == "Test Run"].copy()
        else:
            print(f"❌ 'Note' column not found after renaming in: {filepath}")
            device_stats['status'] = 'error_no_note_column'
            return 0, 0, 0, [], device_stats
    else:
        print(f"❌ Not enough columns to rename 7th column to 'Note' in: {filepath}")
        device_stats['status'] = 'error_insufficient_columns'
        return 0, 0, 0, [], device_stats

    # Insert "Device Name" and "MAC Serial #"
    device_name = raw_df.iloc[0, 1]
    mac_serial = str(raw_df.iloc[1, 0]).replace("DevID: ", "")
    mac_serial = hex_upper(mac_serial)  # Normalize serial number format
    df["Device Name"] = device_name
    df["MAC Serial #"] = mac_serial

    # Update device stats
    device_stats['device_name'] = device_name
    device_stats['device_serial'] = mac_serial
    cols = ["Device Name", "MAC Serial #"] + [col for col in df.columns if col not in ["Device Name", "MAC Serial #"]]
    df = df[cols]

    # Add Enable/Disable columns
    df["Enable"] = df["State"].apply(lambda x: 1 if x == "Enable" else "")
    df["Disable"] = df["State"].apply(lambda x: 1 if x == "Disable" else "")

    # Step 4: Clean up duplicate/missing timestamps before heat detection
    df["Date"] = pd.to_datetime(df["Date"])
    df = df.sort_values("Date").reset_index(drop=True)

    # Capture duplicate rows (keep last, discard others)
    duplicates = df[df.duplicated(subset="Date", keep="last")]

    # Remove duplicates by keeping the last occurrence
    df = df.drop_duplicates(subset="Date", keep="last")

    # Generate a complete timestamp range at 1-minute frequency
    full_range = pd.date_range(start=df["Date"].min(), end=df["Date"].max(), freq="1min")
    df_full = pd.DataFrame({"Date": full_range})

    # Merge to find missing times
    df = pd.merge(df_full, df, on="Date", how="left")

    # Forward fill row values from previous valid row
    fill_cols = ["Device Name", "MAC Serial #", "Enable", "Disable", "Note"]
    df[fill_cols] = df[fill_cols].fillna(method="ffill")

    # Save discarded rows
    discarded = pd.concat([duplicates, df[df.isnull().any(axis=1)]], ignore_index=True)
    df = df.dropna(subset=["State", "Supply Temp/C", "Return Temp/C"]).copy()

    # Continue with filtered data
    df_filtered = df.copy()
    df_filtered["Date"] = pd.to_datetime(df_filtered["Date"])
    supply = df_filtered["Supply Temp/C"].values
    return_temp = df_filtered["Return Temp/C"].values  # <-- needed for the new version1a logic

    # Collect temperature statistics for device status report
    device_stats['test_run_rows'] = len(df_filtered)
    device_stats['supply_min'] = float(df_filtered["Supply Temp/C"].min())
    device_stats['supply_max'] = float(df_filtered["Supply Temp/C"].max())
    device_stats['supply_mean'] = float(df_filtered["Supply Temp/C"].mean())
    device_stats['return_min'] = float(df_filtered["Return Temp/C"].min())
    device_stats['return_max'] = float(df_filtered["Return Temp/C"].max())
    device_stats['return_mean'] = float(df_filtered["Return Temp/C"].mean())

    # Calculate temperature differential stats
    temp_diff = df_filtered["Supply Temp/C"] - df_filtered["Return Temp/C"]
    device_stats['diff_min'] = float(temp_diff.min())
    device_stats['diff_max'] = float(temp_diff.max())
    device_stats['diff_mean'] = float(temp_diff.mean())
    device_stats['rows_above_7c_threshold'] = int((temp_diff >= 7.0).sum())
    device_stats['rows_supply_gt_return'] = int((temp_diff > 0).sum())

    # ----------------------------------------------------------------------------
    # Heating detection logic merged from version1a:
    #   - threshold to turn ON: delta > 5, supply[i] > return_temp[i], not triggered before
    #   - threshold to turn OFF: delta <= -2.7
    #   - maintain a triggered_rows set
    # ----------------------------------------------------------------------------
    heating_state, heating_group = [], []
    in_heating = False
    group_id = 0
    triggered_rows = set()

    for i in range(len(supply)):
        if i == 0:
            heating_state.append("Off")
            heating_group.append(0)
            continue

        delta = supply[i] - supply[i - 1]
        if (not in_heating
            and delta > 5
            and supply[i] > return_temp[i]
            and i not in triggered_rows):
            in_heating = True
            group_id += 1
            heating_state.append("On")
            heating_group.append(group_id)
            triggered_rows.add(i)

        elif in_heating and delta <= -2.7:
            in_heating = False
            heating_state.append("Off")
            heating_group.append(0)

        else:
            heating_state.append("On" if in_heating else "Off")
            heating_group.append(group_id if in_heating else 0)

    df_filtered["Heating"] = heating_state
    df_filtered["Heating_Group"] = heating_group

    # Extra debugging columns (from version1a)
    df_filtered["Delta Supply"] = df_filtered["Supply Temp/C"].diff()
    df_filtered["Delta Debug"] = df_filtered["Delta Supply"].apply(
        lambda x: f"{x:.2f}" if pd.notnull(x) else ""
    )
    df_filtered["Trigger Debug"] = [
        "TRIGGERED" if df_filtered.loc[i, "Delta Supply"] > 5 else ""
        for i in df_filtered.index
    ]

    # ----------------------------------------------------------------------------
    # Validate heating groups (already updated to 0.6 in version2)
    # ----------------------------------------------------------------------------
    valid_groups = []
    total_groups = len([gid for gid in df_filtered["Heating_Group"].unique() if gid > 0])
    for gid in df_filtered["Heating_Group"].unique():
        if gid == 0:
            continue
        group_rows = df_filtered[df_filtered["Heating_Group"] == gid]
        valid_rows = (group_rows["Supply Temp/C"] >= group_rows["Return Temp/C"] + 7).sum()
        if valid_rows / len(group_rows) >= 0.6:
            valid_groups.append(gid)

    # Update device stats with heating group counts
    device_stats['heating_groups_detected'] = total_groups
    device_stats['valid_heating_groups'] = len(valid_groups)

    df_filtered.loc[~df_filtered["Heating_Group"].isin(valid_groups), ["Heating", "Heating_Group"]] = ["Off", 0]

    # Extract heating group on/off times
    group_times = df_filtered[df_filtered["Heating_Group"] > 0].groupby("Heating_Group")["Date"].agg(["min", "max"]).reset_index()
    group_times.columns = ["Heating_Group", "Heat On", "Heat Off"]
    df_filtered = df_filtered.merge(group_times, on="Heating_Group", how="left")

    # Step 6: Keep full hours where heating occurred
    df_filtered["Hour"] = df_filtered["Date"].dt.floor("h")
    df["Date"] = pd.to_datetime(df["Date"])
    df = pd.merge(df, df_filtered[["Date", "Heating", "Heating_Group"]], on="Date", how="left")
    df["Heating"] = df["Heating"].fillna("Off")
    df["Heating_GROUP"] = df["Heating_Group"].fillna(0).astype(int)
    df["Hour"] = df["Date"].dt.floor("h")
    heat_data_set = df[df["Hour"].isin(df_filtered[df_filtered["Heating"] == "On"]["Hour"].unique())].copy()

    # Step 7: Summarize hours with >=55 mins consistent Enable/Disable
    summaries = []
    for hour in sorted(set(heat_data_set["Date"].dt.floor("h"))):
        hour_data = heat_data_set[heat_data_set["Date"].dt.floor("h") == hour]
        enable_count = pd.to_numeric(hour_data["Enable"], errors='coerce').fillna(0).astype(int).sum()
        disable_count = pd.to_numeric(hour_data["Disable"], errors='coerce').fillna(0).astype(int).sum()

        if enable_count >= 55 or disable_count >= 55:
            summaries.append({
                "Device Name": hour_data["Device Name"].iloc[0],
                "MAC Serial #": hour_data["MAC Serial #"].iloc[0],
                "Date/Time On": hour,
                "Date/Time Off": hour + timedelta(minutes=59),
                "Enable": 1 if enable_count >= 55 else "",
                "Disable": 1 if disable_count >= 55 else "",
                "Heating On": (hour_data["Heating"] == "On").sum()
            })

    summary_df = pd.DataFrame(summaries)
    summary_rows_count = len(summary_df)
    print("Summary Rows:", summary_rows_count)

    # Update device stats with final results and status
    device_stats['summary_rows'] = summary_rows_count
    if summary_rows_count > 0:
        device_stats['status'] = 'success'
    elif device_stats['rows_above_7c_threshold'] == 0:
        device_stats['status'] = 'no_heating_detected'
    else:
        device_stats['status'] = 'heating_failed_validation'

    # Save outputs (initial save)
    with pd.ExcelWriter(savepath, engine='openpyxl') as writer:
        original_df.to_excel(writer, sheet_name="Original Data", index=False)
        df.to_excel(writer, sheet_name="Filtered Test Run", index=False)
        heat_data_set.to_excel(writer, sheet_name="Heating Data Set", index=False)
        summary_df.to_excel(writer, sheet_name="Heat Cleaned Data", index=False)
        discarded.to_excel(writer, sheet_name="Discarded", index=False)

    print(f"✅ Processed and saved: {savepath}")

    # ----------------------------------------------------------------------------
    # version1a's highlighting logic
    # ----------------------------------------------------------------------------
    wb = load_workbook(savepath)
    ws = wb["Filtered Test Run"]

    light_orange_fill = PatternFill(start_color="FFD8B1", end_color="FFD8B1", fill_type="solid")

    # Find column letters for relevant headers
    headers = {cell.value: cell.column_letter for cell in ws[1]}
    supply_col = headers.get("Supply Temp/C")
    heating_col = headers.get("Heating")

    # Highlight cells in "Filtered Test Run" where Heating == "On"
    for row in range(2, ws.max_row + 1):
        if ws[f"{heating_col}{row}"].value == "On":
            if supply_col:
                ws[f"{supply_col}{row}"].fill = light_orange_fill
            ws[f"{heating_col}{row}"].fill = light_orange_fill

    wb.save(savepath)

    # ----------------------------------------------------------------------------
    # Insert data into the DB if requested
    # ----------------------------------------------------------------------------
    if insert_db and not summary_df.empty:
        print("ℹ️  Inserting device into 'heating_device'... this may be skipped if dry-run or conflict")
        unique_serials = summary_df["MAC Serial #"].unique()

        if len(unique_serials) != 1:
            print("❌ More than one distinct device_serial found in summary data. Aborting DB insertion.")
            device_stats['status'] = 'error_multiple_serials'
            return summary_rows_count, heating_devices_count, heating_device_readings_count, heating_serial_devices, device_stats

        serial_for_device = str(unique_serials[0])
        serial_for_device = hex_upper(serial_for_device)  # Normalize serial number format
        insert_device_query = """
            INSERT INTO heating_device (device_serial)
            VALUES (%s)
            ON CONFLICT (device_serial) DO NOTHING
            RETURNING device_id, device_serial
        """

        # Execute or dry-run device insert
        if dry_run:
            heating_devices_count = len(unique_serials)
            print(f"Dry run: would execute SQL:\n{insert_device_query.strip()}\nwith parameters ({serial_for_device},)")
            # Add a mock device entry for dry run
            heating_serial_devices.append({
                "device_id": 0,  # Placeholder ID for dry run
                "device_serial": serial_for_device
            })
        else:
            try:
                conn = connect_to_postgres()
                cur = conn.cursor()
                cur.execute(insert_device_query, (serial_for_device,))
                
                # Get the device_id from the RETURNING clause
                device_result = cur.fetchone()
                
                # If a row was returned, a new device was inserted
                # Otherwise, we need to query for the existing device_id
                if device_result:
                    device_id, device_serial = device_result
                    heating_serial_devices.append({
                        "device_id": device_id,
                        "device_serial": device_serial
                    })
                    heating_devices_count = 1
                else:
                    # Query to get the device_id for the existing device
                    cur.execute("SELECT device_id, device_serial FROM heating_device WHERE device_serial = %s", 
                                (serial_for_device,))
                    device_id, device_serial = cur.fetchone()
                    heating_serial_devices.append({
                        "device_id": device_id,
                        "device_serial": device_serial
                    })
                    heating_devices_count = 1
                
            except Exception as e:
                print(f"❌ Failed to insert device into 'heating_device': {e}")
                device_stats['status'] = 'error_db_insertion'
                return summary_rows_count, heating_devices_count, heating_device_readings_count, heating_serial_devices, device_stats

        detroit_tz = pytz.timezone("America/Detroit")
        # Count reading insert attempts
        heating_device_readings_count = len(summary_df)

        for _, row in summary_df.iterrows():
            device_serial = str(row["MAC Serial #"])
            device_serial = hex_upper(device_serial)  # Normalize serial number format

            # Convert local time to epoch UTC
            detroit_time_on = detroit_tz.localize(row["Date/Time On"])
            utc_time_on = detroit_time_on.astimezone(pytz.utc)
            epoch_date_stamp = int(utc_time_on.timestamp())
            date_stamp = utc_time_on.replace(tzinfo=None)

            energy_saver_on = bool(row["Enable"] == 1)
            heating_on_minutes = int(row["Heating On"])
            device_name = str(row["Device Name"])
            date_time_on = row["Date/Time On"]
            date_time_off = row["Date/Time Off"]

            if do_upserts:
                insert_query = """
                    INSERT INTO heating_device_data (
                        device_serial, epoch_date_stamp, date_stamp,
                        energy_saver_on, heating_on_minutes, device_name,
                        date_time_on, date_time_off
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (device_serial, epoch_date_stamp)
                    DO UPDATE SET
                        date_stamp = EXCLUDED.date_stamp,
                        energy_saver_on = EXCLUDED.energy_saver_on,
                        heating_on_minutes = EXCLUDED.heating_on_minutes,
                        device_name = EXCLUDED.device_name,
                        date_time_on = EXCLUDED.date_time_on,
                        date_time_off = EXCLUDED.date_time_off
                    RETURNING device_serial
                """
            else:
                insert_query = """
                    INSERT INTO heating_device_data (
                        device_serial, epoch_date_stamp, date_stamp,
                        energy_saver_on, heating_on_minutes, device_name,
                        date_time_on, date_time_off
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (device_serial, epoch_date_stamp)
                    DO NOTHING
                    RETURNING device_serial
                """

            params = (
                device_serial,
                epoch_date_stamp,
                date_stamp,
                energy_saver_on,
                heating_on_minutes,
                device_name,
                date_time_on,
                date_time_off
            )

            if dry_run:
                print(f"Dry run: would execute SQL:\n{insert_query.strip()}\nwith parameters {params}")
            else:
                cur.execute(insert_query, params)
                # result fetched but we count all attempts uniformly

        if not dry_run:
            conn.commit()
            cur.close()
            conn.close()

        if not do_upserts and heating_device_readings_count > 0:
            print(f"⚠️  Completed inserts with potential duplicates skipped by ON CONFLICT.")
        print("✅ Data insertion complete.")

    # Return counters for JSON summary and device stats
    return summary_rows_count, heating_devices_count, heating_device_readings_count, heating_serial_devices, device_stats


def main():
    """
    Main entry point for the Heat Cycle Detection script.
    - Defaults to current working directory for source_folder.
    - If --input-file is provided, only that file is processed.
    - Otherwise, processes all .xlsx files in the current directory.
    """
    import argparse

    parser = argparse.ArgumentParser(description="Heat Cycle Detection Script")
    parser.add_argument(
        "--input-file",
        help="Process only the specified Excel file (relative to current directory)"
    )
    parser.add_argument(
        "--insert-db",
        action="store_true",
        help="Insert the 'Heat Cleaned Data' rows into the heating_device_data table"
    )
    parser.add_argument(
        "--upserts",
        action="store_true",
        help="Perform upserts (INSERT ... ON CONFLICT DO UPDATE) instead of DO NOTHING"
    )
    parser.add_argument(
        "--logging",
        action="store_true",
        help="Enable logging to files in ./logs"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Display SQL statements without executing inserts/updates"
    )
    args = parser.parse_args()

    # Capture original stdout to ensure JSON summary always goes there
    orig_stdout = sys.stdout

    # Setup logging redirection if requested
    if args.logging:
        now = datetime.now()
        now_str = now.strftime("%Y-%m-%dT%H:%M:%S.%f")[:-4]
        logs_dir = os.path.join(os.getcwd(), "logs")
        os.makedirs(logs_dir, exist_ok=True)
        out_log_path = os.path.join(logs_dir, f"{now_str}.out.log")
        err_log_path = os.path.join(logs_dir, f"{now_str}.err.log")
        out_f = open(out_log_path, 'w')
        err_f = open(err_log_path, 'w')
        sys.stdout = out_f
        sys.stderr = err_f
        print(f"Logging to {out_log_path} (stdout) and {err_log_path} (stderr)", file=sys.stderr)

    # Initialize aggregate counters
    total_summary_rows = 0
    total_heating_devices = 0
    total_heating_readings = 0
    # Initialize list to collect all heating device entries
    all_heating_serial_devices = []
    # Initialize list to collect all device statistics
    all_device_stats = []

    default_source_folder = os.getcwd()

    # Determine upload-results directory (sibling to uploads)
    uploads_dir = os.path.join(default_source_folder, target_folder)
    parent_dir = os.path.dirname(uploads_dir) if target_folder != "." else default_source_folder
    upload_results_dir = os.path.join(parent_dir, "upload-results")
    os.makedirs(upload_results_dir, exist_ok=True)

    if args.input_file:
        input_path = os.path.join(default_source_folder, args.input_file)
        if os.path.isfile(input_path) and input_path.endswith(".xlsx") and not os.path.basename(input_path).startswith("~$"):
            name, _ = os.path.splitext(os.path.basename(input_path))
            save_path = os.path.join(target_folder, f"{name}_heat min per hour.xlsx")
            summary, dev_count, read_count, devices, stats = process_file(
                input_path,
                save_path,
                insert_db=args.insert_db,
                do_upserts=args.upserts,
                dry_run=args.dry_run
            )
            total_summary_rows += summary
            total_heating_devices += dev_count
            total_heating_readings += read_count
            all_heating_serial_devices.extend(devices)
            all_device_stats.append(stats)
        else:
            print(f"❌ File not found or invalid format: {input_path}")
    else:
        for filename in os.listdir(default_source_folder):
            if filename.endswith(".xlsx") and not filename.startswith("~$"):
                full_path = os.path.join(default_source_folder, filename)
                name, _ = os.path.splitext(filename)
                save_path = os.path.join(target_folder, f"{name}_heat min per hour.xlsx")
                summary, dev_count, read_count, devices, stats = process_file(
                    full_path,
                    save_path,
                    insert_db=args.insert_db,
                    do_upserts=args.upserts,
                    dry_run=args.dry_run
                )
                total_summary_rows += summary
                total_heating_devices += dev_count
                total_heating_readings += read_count
                all_heating_serial_devices.extend(devices)
                all_device_stats.append(stats)

    # Generate device-status.xlsx report if any files were processed
    if all_device_stats:
        # Reorder columns for better readability
        column_order = [
            'filepath', 'device_name', 'device_serial', 'status',
            'test_run_rows', 'summary_rows',
            'supply_min', 'supply_max', 'supply_mean',
            'return_min', 'return_max', 'return_mean',
            'diff_min', 'diff_max', 'diff_mean',
            'rows_supply_gt_return', 'rows_above_7c_threshold',
            'heating_groups_detected', 'valid_heating_groups'
        ]

        # Create vertical format: headers in column A, values in column B, comments in column C
        # With blank rows between multiple records
        vertical_data = []
        for idx, stats in enumerate(all_device_stats):
            # Add blank row separator between records (except before first)
            if idx > 0:
                vertical_data.append(['', '', ''])

            # Add each field as a row with explanatory comments
            for col in column_order:
                value = stats.get(col, '')
                comment = ''

                # Add comments for threshold-related fields
                if col == 'status':
                    if value == 'no_heating_detected':
                        comment = 'NO HEATING: Supply temp never exceeded return by required 7°C'
                    elif value == 'heating_failed_validation':
                        comment = 'VALIDATION FAILED: Heating detected but could not sustain 7°C for 60% of cycle'
                    elif value == 'success':
                        comment = 'SUCCESS: Valid heating cycles detected and processed'

                elif col == 'summary_rows':
                    if value == 0:
                        comment = 'ZERO OUTPUT: No valid heating hours generated for database insertion'

                elif col == 'diff_max':
                    if isinstance(value, (int, float)):
                        if value < 7.0:
                            comment = f'BELOW THRESHOLD: Max differential {value:.1f}°C < 7.0°C required for validation'
                        elif value >= 7.0:
                            comment = f'Above 7°C threshold (validation requirement met for this metric)'

                elif col == 'diff_mean':
                    if isinstance(value, (int, float)):
                        if value < 0:
                            comment = 'NEGATIVE: Supply cooler than return (indicates cooling mode or sensor issue)'
                        elif value >= 0 and value < 7.0:
                            comment = 'POSITIVE but below 7°C: Some heating activity but insufficient for validation'

                elif col == 'rows_above_7c_threshold':
                    if value == 0:
                        comment = 'CRITICAL: No rows met 7°C validation threshold - cannot generate valid heating cycles'
                    elif isinstance(value, int) and value > 0:
                        total_rows = stats.get('test_run_rows', 1)
                        pct = (value / total_rows * 100) if total_rows > 0 else 0
                        comment = f'{pct:.1f}% of total rows meet 7°C threshold'

                elif col == 'valid_heating_groups':
                    detected = stats.get('heating_groups_detected', 0)
                    if value == 0 and detected > 0:
                        comment = f'VALIDATION FAILED: {detected} groups detected but none sustained 7°C for 60% of cycle'
                    elif value > 0:
                        comment = f'{value}/{detected} detected groups passed 60% validation rule'

                elif col == 'heating_groups_detected':
                    if value == 0:
                        comment = 'No temperature patterns triggered heating detection (>5°C jump + supply>return)'
                    elif value > 0:
                        comment = f'{value} potential heating cycles detected (requires validation)'

                vertical_data.append([col, value, comment])

        # Create DataFrame with vertical format including comments
        vertical_df = pd.DataFrame(vertical_data, columns=['Field', 'Value', 'Comment'])

        # Determine output filename based on input mode
        if args.input_file:
            # Single file mode: use input filename with -results.xlsx
            input_basename = os.path.basename(args.input_file)
            input_name_no_ext = os.path.splitext(input_basename)[0]
            status_filename = f"{input_name_no_ext}-results.xlsx"
        else:
            # Batch mode: use timestamp-based filename
            now = datetime.now()
            timestamp = now.strftime("%Y%m%d_%H%M%S")
            status_filename = f"batch-{timestamp}-results.xlsx"

        # Write to upload-results directory
        status_report_path = os.path.join(upload_results_dir, status_filename)
        vertical_df.to_excel(status_report_path, index=False, engine='openpyxl')

        # Format the Excel file for better readability
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment

        wb = load_workbook(status_report_path)
        ws = wb.active

        # Make Field column (column A) bold
        bold_font = Font(bold=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.font = bold_font

        # Left-justify Value column (column B) - all data rows, not header
        left_align = Alignment(horizontal='left')
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.alignment = left_align

        # Auto-size all columns to fit content
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    cell_value = str(cell.value) if cell.value is not None else ''
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass

            # Add some padding and set column width
            adjusted_width = min(max_length + 2, 100)  # Cap at 100 to avoid extremely wide columns
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(status_report_path)
        print(f"📊 Device status report written to: {status_report_path}")

    # Output JSON summary regardless of logging
    mode = "dry-run" if args.dry_run else "live-run"
    summary_obj = {
        "mode": mode,
        "summary-rows": total_summary_rows,
        "heating-devices": total_heating_devices,
        "heating-device-readings": total_heating_readings,
        "heating-serial-devices": all_heating_serial_devices
    }
    orig_stdout.write(json.dumps(summary_obj) + "\n")

if __name__ == "__main__":
    main()
